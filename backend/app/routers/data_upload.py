import io
import re
import logging
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.database import get_db
from app.dependencies import get_current_user

router = APIRouter()
logger = logging.getLogger(__name__)

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]


def _parse_overtime_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))

    overtime_row = None
    working_row = None
    for row in rows:
        if row is None:
            continue
        type_val = str(row[1]).strip().lower() if row[1] is not None else ''
        if 'overtime hour' in type_val:
            overtime_row = row
        elif 'working hour' in type_val:
            working_row = row

    if overtime_row is None or working_row is None:
        raise HTTPException(
            status_code=422,
            detail="Format tidak valid: baris 'Overtime Hour' atau 'Working Hour' tidak ditemukan. "
                   "Pastikan kolom B berisi label tersebut.",
        )

    result = []
    for i, month_name in enumerate(MONTH_NAMES):
        ot = float(overtime_row[i + 2] or 0) if (i + 2) < len(overtime_row) else 0.0
        wk = float(working_row[i + 2] or 0) if (i + 2) < len(working_row) else 0.0
        total = ot + wk
        ratio = round(ot / total * 100, 2) if total > 0 else 0.0
        result.append({
            'period_num': i + 1,
            'period_name': month_name,
            'overtime_hours': round(ot, 2),
            'working_hours': round(wk, 2),
            'ratio_pct': ratio,
        })
    return result


@router.get("/overtime")
async def get_overtime_data(
    year: int = Query(2025),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Return uploaded overtime data for the given year."""
    q = text("""
        SELECT per.period_num, per.period_name,
               o.overtime_hours, o.working_hours,
               CASE WHEN (o.working_hours + o.overtime_hours) > 0
                    THEN ROUND((o.overtime_hours / (o.working_hours + o.overtime_hours) * 100)::numeric, 2)
                    ELSE 0 END AS ratio_pct
        FROM eis.fact_overtime o
        JOIN eis.dim_period per ON o.period_id = per.id
        WHERE per.fiscal_year = :year
        ORDER BY per.period_num
    """)
    result = await db.execute(q, {"year": year})
    return {"data": [dict(r) for r in result.mappings().all()]}


@router.post("/overtime/upload")
async def upload_overtime(
    year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Parse overtime Excel and upsert into fact_overtime."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="File harus .xlsx atau .xls")

    content = await file.read()
    months = _parse_overtime_excel(content)

    loaded = 0
    for m in months:
        # Find period_id
        res = await db.execute(
            text("SELECT id FROM eis.dim_period WHERE fiscal_year=:y AND period_num=:p"),
            {"y": year, "p": m["period_num"]},
        )
        row = res.fetchone()
        if not row:
            continue
        period_id = row[0]

        await db.execute(text("""
            INSERT INTO eis.fact_overtime (period_id, overtime_hours, working_hours)
            VALUES (:pid, :ot, :wk)
            ON CONFLICT (period_id) DO UPDATE SET
                overtime_hours = EXCLUDED.overtime_hours,
                working_hours  = EXCLUDED.working_hours
        """), {"pid": period_id, "ot": m["overtime_hours"], "wk": m["working_hours"]})
        loaded += 1

    await db.commit()
    return {
        "message": f"Berhasil upload {loaded} bulan overtime untuk tahun {year}",
        "year": year,
        "loaded": loaded,
        "data": months,
    }


# ══════════════════════════════════════════════════════════════
# COGS Upload — format baru (COGS Data New.xlsx)
#
# Struktur Excel:
#   Row 1-2 : judul / catatan
#   Row 3   : header utama (No, Market, Customer, Products,
#             Price USD, Price IDR, COGS Quantity, COGS Amount)
#   Row 4   : sub-header (tahun untuk Price, nama bulan Jan-Dec)
#   Row 5+  : data produk
#
# Kolom (0-indexed):
#   col 1  = No (nomor urut)
#   col 2  = Market  (Public / Private / CMO / Export / Service Agreement)
#   col 3  = Customer
#   col 4  = Products (nama produk pendek + berat, e.g. "Carboplatin 150 mg")
#   col 6  = Price USD 2024
#   col 7  = Price USD 2025
#   col 7  = Price IDR 2024
#   col 8  = Price IDR 2025
#   col 9–20  = COGS Quantity per bulan Jan–Dec
#   col 21 = Total Quantity
#   col 22–33 = COGS Amount per bulan Jan–Dec (juta IDR)
#   col 34 = Total COGS Amount
#   col 37 = Nama produk lengkap dengan brand
#   col 40 = Kode barang (opsional, diisi jika ada)
# ══════════════════════════════════════════════════════════════

_MARKET_BIZ_TYPE = {
    "public":            "Local",
    "private":           "Local",
    "service agreement": "Local",
    "accounting":        "Local",
    "cmo":               "CMO",
    "export":            "Export",
}

_MARKET_CODE = {
    "public":            "PUB",
    "private":           "PRI",
    "cmo":               "CMO",
    "export":            "EXP",
    "service agreement": "SVC",
    "accounting":        "ACC",
}


def _extract_base_name(product_name: str) -> str:
    """Hapus suffix kekuatan/berat dari nama produk.

    Contoh:
      'Carboplatin 150 mg'      → 'Carboplatin'
      'Paclitaxel 30 mg'        → 'Paclitaxel'
      'Gemcitabine 200 mg - Liquid' → 'Gemcitabine'
      'Darbepoetin alfa 20 mcg' → 'Darbepoetin alfa'
    """
    base = re.sub(
        r'\s+[\d,\.]+\s*(mg|mcg|gr|g(?!em)|ml|mL|L|IU|iu|µg|μg)(/\w+)?\s*$',
        '', product_name, flags=re.IGNORECASE
    ).strip()
    base = re.sub(r'\s*[-–]\s*Liquid\s*$', '', base, flags=re.IGNORECASE).strip()
    return base if base else product_name


def _make_product_code(base_name: str, market: str) -> str:
    """Generate product_code <= 20 chars: {BASECODE}_{MKT}."""
    mkt_code = _MARKET_CODE.get(market.lower(), "OTH")
    clean    = re.sub(r'[^A-Za-z0-9]', '', base_name)[:12].upper()
    return f"{clean}_{mkt_code}"


def _parse_cogs_new_excel(content: bytes) -> list[dict]:
    """Parse COGS Data New.xlsx dan gabungkan per (base_name, market).

    Produk dengan kekuatan berbeda (mis. Paclitaxel 30 mg, 100 mg, 300 mg)
    digabungkan menjadi satu entri 'Paclitaxel' per market.
    Quantity dan COGS Amount dijumlahkan per bulan.

    Returns list of dicts:
      {product_code, display_name, biz_type, market,
       price_idr (weighted avg), qty[12], cogs_amt[12]}
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    groups: dict = {}

    for row in ws.iter_rows(min_row=5, values_only=True):
        no = row[1] if len(row) > 1 else None
        if no is None or not isinstance(no, (int, float)):
            continue

        market        = str(row[2] or "Public").strip()
        product_short = str(row[4] or "").strip()
        price_idr     = float(row[8] or 0) if len(row) > 8 else 0.0

        qty      = [float(row[9  + i] or 0) if len(row) > 9  + i else 0.0 for i in range(12)]
        cogs_amt = [float(row[22 + i] or 0) if len(row) > 22 + i else 0.0 for i in range(12)]

        if not any(qty) and not any(cogs_amt):
            continue

        base_name = _extract_base_name(product_short)
        key       = (base_name, market)

        if key not in groups:
            groups[key] = {
                "product_code": _make_product_code(base_name, market)[:20],
                "display_name": base_name[:150],
                "biz_type":     _MARKET_BIZ_TYPE.get(market.lower(), "Local"),
                "market":       market,
                "price_idr":    price_idr,
                "qty":          [0.0] * 12,
                "cogs_amt":     [0.0] * 12,
                "_total_qty":   0.0,
                "_total_sales": 0.0,
            }

        g = groups[key]
        for i in range(12):
            g["qty"][i]      += qty[i]
            g["cogs_amt"][i] += cogs_amt[i]

        row_qty = sum(qty)
        g["_total_qty"]   += row_qty
        g["_total_sales"] += price_idr * row_qty

    if not groups:
        raise HTTPException(
            status_code=422,
            detail=(
                "Tidak ada data yang bisa dibaca. "
                "Pastikan format file sesuai: baris 3 header, baris 4 nama bulan, "
                "baris 5+ data produk dengan kolom No di kolom B."
            ),
        )

    # Finalize weighted-average price per group
    result = []
    for g in groups.values():
        if g["_total_qty"] > 0:
            g["price_idr"] = g["_total_sales"] / g["_total_qty"]
        del g["_total_qty"], g["_total_sales"]
        result.append(g)

    return result


@router.get("/cogs")
async def get_cogs_data(
    year: int = Query(2025),
    period: int = Query(12),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Return COGS ratio data for the given year/period (from fact_cogs)."""
    q = text("""
        SELECT p.product_name,
               ROUND(SUM(c.sales_amount)::numeric, 2)  AS net_sales,
               ROUND(SUM(c.cogs_total)::numeric, 2)    AS cogs,
               CASE WHEN SUM(c.sales_amount) > 0
                    THEN ROUND((SUM(c.cogs_total) / SUM(c.sales_amount) * 100)::numeric, 2)
                    ELSE 0 END                          AS cogs_pct
        FROM eis.fact_cogs c
        JOIN eis.dim_product p   ON c.product_id = p.id
        JOIN eis.dim_period  per ON c.period_id  = per.id
        WHERE per.fiscal_year = :year AND per.period_num <= :period
        GROUP BY p.product_name
        HAVING SUM(c.sales_amount) > 0
        ORDER BY cogs_pct DESC
    """)
    result = await db.execute(q, {"year": year, "period": period})
    return {"data": [dict(r) for r in result.mappings().all()]}


@router.post("/cogs/upload")
async def upload_cogs(
    year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Parse COGS Excel (format baru) dan insert langsung per bulan ke fact_cogs.

    Format file: COGS Data New.xlsx
    - Baris 3   : header kolom
    - Baris 4   : nama bulan (Jan–Dec)
    - Baris 5+  : data produk
    - Kolom B   : No urut
    - Kolom C   : Market (Public/Private/CMO/Export/Service Agreement)
    - Kolom E   : Nama produk (pendek)
    - Kolom I   : Price IDR tahun upload
    - Kolom J–U : COGS Quantity Jan–Dec
    - Kolom W–AH: COGS Amount Jan–Dec (juta IDR)
    - Kolom AL  : Nama produk lengkap dengan brand (opsional)
    - Kolom AO  : Kode barang (opsional, jika kosong digenerate otomatis)
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="File harus .xlsx atau .xls")

    content = await file.read()
    records = _parse_cogs_new_excel(content)

    # Build period_id lookup for the target year
    periods_res = await db.execute(
        text("SELECT period_num, id FROM eis.dim_period WHERE fiscal_year=:y ORDER BY period_num"),
        {"y": year},
    )
    period_map = {r[0]: r[1] for r in periods_res.fetchall()}

    if not period_map:
        raise HTTPException(status_code=422, detail=f"Tidak ada data periode untuk tahun {year}")

    loaded_products = 0
    loaded_cogs = 0
    skipped = []
    results = []

    for rec in records:
        # Upsert dim_product
        await db.execute(text("""
            INSERT INTO eis.dim_product (product_code, product_name, business_type, market)
            VALUES (:code, :name, :biz, :mkt)
            ON CONFLICT (product_code) DO UPDATE SET
                product_name  = EXCLUDED.product_name,
                business_type = EXCLUDED.business_type,
                market        = EXCLUDED.market
        """), {
            "code": rec["product_code"],
            "name": rec["display_name"],
            "biz":  rec["biz_type"],
            "mkt":  rec["market"],
        })

        prod_res = await db.execute(
            text("SELECT id FROM eis.dim_product WHERE product_code = :code"),
            {"code": rec["product_code"]},
        )
        product_id = prod_res.fetchone()[0]
        loaded_products += 1

        month_loaded = 0
        for month_idx, (qty, cogs_amt) in enumerate(zip(rec["qty"], rec["cogs_amt"])):
            if qty == 0 and cogs_amt == 0:
                continue

            month_num = month_idx + 1
            period_id = period_map.get(month_num)
            if not period_id:
                continue

            # sales_amount = Price IDR × Qty ÷ 1,000,000 (juta IDR)
            sales_amt = round(rec["price_idr"] * qty / 1_000_000, 4)
            ebit_amt  = round(sales_amt - cogs_amt, 4)

            await db.execute(text("""
                INSERT INTO eis.fact_cogs
                    (period_id, product_id, sales_amount, cogs_total, ebit_amount)
                VALUES (:pid, :prod_id, :sales, :cogs, :ebit)
                ON CONFLICT (period_id, product_id) DO UPDATE SET
                    sales_amount = EXCLUDED.sales_amount,
                    cogs_total   = EXCLUDED.cogs_total,
                    ebit_amount  = EXCLUDED.ebit_amount
            """), {
                "pid":     period_id,
                "prod_id": product_id,
                "sales":   sales_amt,
                "cogs":    round(cogs_amt, 4),
                "ebit":    ebit_amt,
            })
            month_loaded += 1
            loaded_cogs += 1

        results.append({
            "product_code":  rec["product_code"],
            "product_name":  rec["display_name"],
            "market":        rec["market"],
            "months_loaded": month_loaded,
        })
        if month_loaded == 0:
            skipped.append(rec["product_code"])

    await db.commit()
    logger.info(f"[upload_cogs] {loaded_products} products, {loaded_cogs} cogs rows for {year}")

    return {
        "message": f"Berhasil upload {loaded_products} produk ({loaded_cogs} baris COGS) untuk tahun {year}",
        "year":    year,
        "products": loaded_products,
        "cogs_rows": loaded_cogs,
        "skipped": skipped,
        "data":    results,
    }


# ══════════════════════════════════════════════════════════════
# SALES BP Upload — format DATA BP.xlsx
#
# Format Excel yang didukung:
#   Row header bulan (Jan–Dec) diikuti baris segment:
#     Total  | 800 | 900 | ... | 880 |
#     Local  | 500 | 600 | ... | 550 |
#     CMO    | 200 | 200 | ... | 220 |
#     Export | 100 | 100 | ... | 110 |
#   (atau format matriks serupa, label di kolom pertama)
#
# Nilai dalam juta IDR.
# ══════════════════════════════════════════════════════════════

_SEGMENT_MAP = {
    'local': 'Local', 'lokal': 'Local', 'total local': 'Local', 'total lokal': 'Local',
    'cmo': 'CMO', 'total cmo': 'CMO',
    'export': 'Export', 'ekspor': 'Export', 'total export': 'Export', 'total ekspor': 'Export',
    'total': 'Total', 'grand total': 'Total', 'jumlah': 'Total',
}

_BP_MONTH_SHORT = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
_BP_MONTH_LONG  = ['january', 'february', 'march', 'april', 'may', 'june',
                   'july', 'august', 'september', 'october', 'november', 'december']


def _parse_sales_bp_excel(content: bytes) -> dict:
    """Parse DATA BP.xlsx untuk mendapatkan monthly sales BP per segment.

    Mendukung dua strategi:
    A) Row header bulan ditemukan → baca baris setelahnya berdasarkan label segment.
    B) Fallback: scan baris yang mengandung label segment + 12 angka berurutan.

    Returns:
        {
            'segments': {'Total': [12 floats], 'Local': [...], 'CMO': [...], 'Export': [...]},
            'detected_year': int | None,
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Deteksi tahun dari 5 baris pertama
    detected_year = None
    for row in all_rows[:5]:
        for v in (row or []):
            if isinstance(v, (int, float)) and 2020 <= v <= 2030:
                detected_year = int(v)
                break
        if detected_year:
            break

    # Strategi A: cari header row yang berisi nama bulan
    header_row_idx = None
    month_col_map: dict[int, int] = {}

    for row_idx, row in enumerate(all_rows):
        if not row:
            continue
        row_lower = [str(v).strip().lower() if v is not None else '' for v in row]
        found: list[tuple[int, int]] = []
        for col_idx, cell in enumerate(row_lower):
            for m_idx, (sh, lo) in enumerate(zip(_BP_MONTH_SHORT, _BP_MONTH_LONG)):
                if cell in (sh, lo):
                    found.append((m_idx, col_idx))
                    break
        if len(found) >= 6:
            header_row_idx = row_idx
            for m_idx, col_idx in found:
                month_col_map[m_idx] = col_idx
            break

    segments: dict[str, list[float]] = {}

    if header_row_idx is not None and len(month_col_map) >= 10:
        for row in all_rows[header_row_idx + 1:]:
            if not row:
                continue
            label = None
            for v in list(row)[:6]:
                if v is not None:
                    s = str(v).strip().lower()
                    if s in _SEGMENT_MAP:
                        label = _SEGMENT_MAP[s]
                        break
            if not label:
                continue

            monthly: list[float] = []
            for m_idx in range(12):
                col_idx = month_col_map.get(m_idx)
                val = row[col_idx] if col_idx is not None and col_idx < len(row) else None
                monthly.append(float(val) if isinstance(val, (int, float)) else 0.0)

            if any(v > 0 for v in monthly):
                segments[label] = monthly
    else:
        # Strategi B: scan setiap baris
        for row in all_rows:
            if not row:
                continue
            for i, v in enumerate(row):
                if v is None:
                    continue
                s = str(v).strip().lower()
                if s in _SEGMENT_MAP:
                    label = _SEGMENT_MAP[s]
                    nums = [x for x in row[i + 1:] if isinstance(x, (int, float))]
                    if len(nums) >= 12:
                        segments[label] = [float(x) for x in nums[:12]]
                    break

    if not segments:
        raise HTTPException(
            status_code=422,
            detail=(
                "Format file tidak dapat dibaca. "
                "Pastikan file memiliki header bulan (Jan–Dec) dan baris segment "
                "(Total / Local / CMO / Export). "
                "Nilai dalam juta IDR."
            ),
        )

    return {'segments': segments, 'detected_year': detected_year}


@router.get("/sales-bp")
async def get_sales_bp(
    year: int = Query(2025),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Return Sales BP per segment per month dari tabel business_plan."""
    q = text("""
        SELECT category,
               jan, feb, mar, apr, may, jun,
               jul, aug, sep, oct, nov, "dec", total
        FROM eis.business_plan
        WHERE fiscal_year = :year AND plan_type = 'Sales'
        ORDER BY
            CASE category
                WHEN 'Total'  THEN 0
                WHEN 'Local'  THEN 1
                WHEN 'CMO'    THEN 2
                WHEN 'Export' THEN 3
                ELSE 4 END
    """)
    result = await db.execute(q, {"year": year})
    return {"data": [dict(r) for r in result.mappings().all()]}


@router.post("/sales-bp/upload")
async def upload_sales_bp(
    year: int = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Parse DATA BP.xlsx, simpan ke business_plan, dan update fact_sales.bp_amount.

    Segmen yang didukung: Total, Local, CMO, Export.
    Nilai dalam juta IDR.

    Setelah upload:
    - Tabel business_plan di-update (referensi untuk ETL berikutnya).
    - fact_sales.bp_amount langsung di-update per bulan per segmen.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="File harus .xlsx atau .xls")

    content = await file.read()
    parsed   = _parse_sales_bp_excel(content)
    segments = parsed['segments']

    created_by = user.get('name', user.get('username', 'upload'))
    loaded_bp     = 0
    updated_sales = 0

    for category, monthly_vals in segments.items():
        # ── Upsert business_plan ────────────────────────────────────
        # Hapus dulu karena UNIQUE constraint tidak bisa pakai ON CONFLICT
        # dengan sub_category IS NULL (NULL != NULL di Postgres).
        await db.execute(text("""
            DELETE FROM eis.business_plan
            WHERE fiscal_year = :year
              AND plan_type   = 'Sales'
              AND category    = :cat
              AND sub_category IS NULL
        """), {"year": year, "cat": category})

        m = monthly_vals
        await db.execute(text("""
            INSERT INTO eis.business_plan
                (fiscal_year, plan_type, category, sub_category,
                 jan, feb, mar, apr, may, jun, jul, aug, sep, oct, nov, "dec",
                 created_by, updated_at)
            VALUES
                (:year, 'Sales', :cat, NULL,
                 :m0,:m1,:m2,:m3,:m4,:m5,:m6,:m7,:m8,:m9,:m10,:m11,
                 :created_by, NOW())
        """), {
            "year": year, "cat": category, "created_by": created_by,
            "m0": m[0],  "m1": m[1],  "m2": m[2],  "m3": m[3],
            "m4": m[4],  "m5": m[5],  "m6": m[6],  "m7": m[7],
            "m8": m[8],  "m9": m[9],  "m10": m[10], "m11": m[11],
        })
        loaded_bp += 1

        # ── Update fact_sales.bp_amount langsung ────────────────────
        if category in ('Local', 'CMO', 'Export'):
            for month_idx, bp_val in enumerate(monthly_vals):
                period_res = await db.execute(
                    text("SELECT id FROM eis.dim_period WHERE fiscal_year=:y AND period_num=:p"),
                    {"y": year, "p": month_idx + 1},
                )
                period_row = period_res.fetchone()
                if not period_row:
                    continue
                period_id = period_row[0]

                # Coba UPDATE dulu
                upd = await db.execute(text("""
                    UPDATE eis.fact_sales
                       SET bp_amount  = :bp,
                           updated_at = NOW()
                     WHERE period_id    = :pid
                       AND business_type = :biz
                       AND market       = 'All'
                       AND product_id IS NULL
                """), {"bp": round(bp_val, 2), "pid": period_id, "biz": category})

                if upd.rowcount == 0:
                    # Belum ada baris ETL → insert baru
                    await db.execute(text("""
                        INSERT INTO eis.fact_sales
                            (period_id, product_id, business_type, market, bp_amount, actual_amount)
                        VALUES (:pid, NULL, :biz, 'All', :bp, 0)
                    """), {"pid": period_id, "biz": category, "bp": round(bp_val, 2)})

                updated_sales += 1

    await db.commit()
    logger.info(
        f"[upload_sales_bp] {loaded_bp} segments, {updated_sales} fact_sales rows, year={year}"
    )

    return {
        "message": (
            f"Berhasil upload Business Plan: {loaded_bp} segmen, "
            f"{updated_sales} baris sales diupdate untuk tahun {year}"
        ),
        "year":               year,
        "loaded_segments":    loaded_bp,
        "updated_sales_rows": updated_sales,
        "detected_year":      parsed.get("detected_year"),
        "data": {cat: vals for cat, vals in segments.items()},
    }
