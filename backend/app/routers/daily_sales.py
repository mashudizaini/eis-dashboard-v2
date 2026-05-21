import json
import io
import datetime
from pathlib import Path
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from app.dependencies import get_current_user

router = APIRouter()

DATA_FILE = Path(__file__).parent.parent / "data" / "daily_sales.json"
DEFAULT_FILE = Path(__file__).parent.parent / "data" / "daily_sales_default.json"

MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december"
]


def _load_data() -> dict:
    for path in (DATA_FILE, DEFAULT_FILE):
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                continue
    return {}


def _save_data(data: dict):
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def _parse_excel(content: bytes) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    chart_idx = next((i for i, s in enumerate(sheet_names_lower) if "chart" in s), None)
    perf_idx = next((i for i, s in enumerate(sheet_names_lower) if "daily sales" in s or "performance" in s), None)

    if chart_idx is None:
        raise HTTPException(status_code=422, detail="Sheet 'Chart' tidak ditemukan di file Excel")

    ws = wb[wb.sheetnames[chart_idx]]
    rows = list(ws.iter_rows(min_row=1, max_row=35, values_only=True))

    month_starts = [2 + i * 3 for i in range(12)]
    table_rows = []
    for row in rows[2:]:
        if row[1] is None or not isinstance(row[1], (int, float)):
            continue
        wd = int(row[1])
        entry = {"wd": wd}
        for i, month in enumerate(MONTHS):
            ms = month_starts[i]
            entry[month] = {
                "target": round(float(row[ms]), 3) if row[ms] is not None else None,
                "acc": round(float(row[ms + 1]), 3) if row[ms + 1] is not None else None,
                "sales": round(float(row[ms + 2]), 3) if row[ms + 2] is not None else None,
            }
        table_rows.append(entry)

    month_targets = {}
    for i, month in enumerate(MONTHS):
        ms = month_starts[i]
        for row in rows[2:]:
            if row[ms] is not None:
                month_targets[month] = round(float(row[ms]), 3)
                break

    bp = 0.0
    exp_closing = 0.0
    ach_pct = 0.0
    as_of = ""
    year = 2025
    month_label = "December"

    if perf_idx is not None:
        ws2 = wb[wb.sheetnames[perf_idx]]
        for row in ws2.iter_rows(min_row=1, max_row=12, values_only=True):
            nums = [v for v in row if isinstance(v, (int, float))]
            if len(nums) >= 3:
                bp_c, exp_c, ach_c = nums[0], nums[1], nums[2]
                if 1000 < bp_c < 50000 and 0 < ach_c < 2:
                    bp = round(bp_c, 3)
                    exp_closing = round(exp_c, 3)
                    ach_pct = round(ach_c * 100, 2)
            dates = [v for v in row if isinstance(v, datetime.datetime)]
            if dates:
                as_of = dates[0].strftime("%Y-%m-%d")
                year = dates[0].year

    return {
        "year": year,
        "month": month_label,
        "as_of": as_of,
        "business_plan": bp,
        "expectation_closing": exp_closing,
        "achievement_pct": ach_pct,
        "month_targets": month_targets,
        "rows": table_rows,
    }


@router.get("/data")
async def get_daily_sales(user: dict = Depends(get_current_user)):
    return {"data": _load_data()}


@router.post("/upload")
async def upload_daily_sales(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=422, detail="File harus berformat .xlsx atau .xls")
    content = await file.read()
    data = _parse_excel(content)
    _save_data(data)
    return {"message": "Data berhasil diupload", "data": data}
